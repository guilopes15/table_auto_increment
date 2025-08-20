import tempfile
from pathlib import Path

import pytest
from openpyxl import Workbook


@pytest.fixture
def temp_path():
    with tempfile.TemporaryDirectory() as tmpdir:
        path = Path(tmpdir) / 'test.xlsx'
        yield path


@pytest.fixture
def temp_table(temp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Produtos'
    ws['A1'] = 'Nome'
    ws['B1'] = 'Quantidade'
    ws['C1'] = 'Descrição'
    wb.save(temp_path)
