from openpyxl import load_workbook

from src.table_auto_increment.schemas import Product
from src.table_auto_increment.table_manager import save_product


def test_append_table_not_exist(temp_path):
    data = Product(name='test', quantity=1, description='apenas um test')
    product = save_product(path=temp_path, data=data)

    assert product


def test_append_table_exist(temp_path, temp_table):
    data = Product(name='test', quantity=1, description='apenas um test')
    product = save_product(path=temp_path, data=data)

    assert product


def test_verify_table_data(temp_path, temp_table):
    data = Product(name='test', quantity=1, description='apenas um test')
    save_product(path=temp_path, data=data)

    wb = load_workbook(temp_path)
    ws = wb['Produtos']

    # ws[2] retorna a segunda linha inteira da tabela
    name, quantity, description = [items.value for items in ws[2]]

    assert name == data.name
    assert quantity == data.quantity
    assert description == data.description
