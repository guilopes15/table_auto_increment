from pathlib import Path

from openpyxl import Workbook, load_workbook

from .schemas import Product


def save_product(path: Path, data: Product):
    if path.exists():
        wb = load_workbook(path)
        ws = wb['Produtos']

    else:
        wb = Workbook()
        ws = wb.active
        ws.title = 'Produtos'
        ws['A1'] = 'Nome'
        ws['B1'] = 'Quantidade'
        ws['C1'] = 'Descrição'

    data = [value for _, value in data.model_dump().items()]
    ws.append(data)
    wb.save(path)

    return path
